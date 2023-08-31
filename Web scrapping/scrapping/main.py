import requests
from requests.auth import HTTPBasicAuth
import json
import openpyxl
import time

url = "https://mygovernmentonline.atlassian.net/gateway/api/graphql"
url1 = "https://mygovernmentonline.atlassian.net/rest/api/2/search"

mail = "ai@scpdc.org"
api_Token = "ATATT3xFfGF0_Zb3KWw7AdSB_7YbkDGV3pTxJ9TgMLQU0Xs7rTgsRno6HyvOYhEfpfUpzAn4ySxw0T_3y9F9d1gpGdsd_owIdP4ykfC4W7zUPwwhuh7ztoBi_v18x4VG2kICW5-hRaThbwoBSI8siTHJYpXhpSvBO_QFbGdFSHZFHnOWIAImH5U=C68B93F3"

auth = HTTPBasicAuth(mail, api_Token)

headers = {
  "Accept": "application/json",
  "X-ExperimentalApi": "JiraIssue, projectStyle",
}

jql = 'project = MWS AND assignee IS EMPTY AND status in (ACKNOWLEDGED, REQUESTED)'

# Set fields to return in the response
fields = 'summary,status,created'

unassigned_headers = {
    'Content-Type': 'application/json'
}

unassigned_params = {
    'jql': jql,
    'fields': 'key',
    'maxResults': 1000,  # Set the maximum number of results per page to 1000
    'startAt': 0,
}

# Make the API call
response1 = requests.get(url1, headers=unassigned_headers, params=unassigned_params, auth=auth)

if response1.status_code != 200:
    print('Error accessing Jira API:', response1.text)
else:
    response_json = response1.json()
    # extract the issue keys from the response json
    issue_keys = [issue['key'] for issue in response_json['issues']]
    # print('The issue keys for unassigned issues:', issue_keys)

query = """
  query mainIssueAggQuery($issueKey:String!,$cloudId:ID!,$issueViewRelayLeaderFlag:Boolean!,$issueViewRelayLabelsFlag:Boolean!,$giraAGGMigrationJiraSettingsFlag:Boolean!,$giraAGGMigrationIsArchivedFlag:Boolean!){jira{epicLinkFieldKey,isSubtasksEnabled(cloudId:$cloudId)@include(if:$giraAGGMigrationJiraSettingsFlag)@optIn(to:\"JiraIsSubtasksEnabled\")\n    globalTimeTrackingSettings(cloudId:$cloudId)@include(if:$giraAGGMigrationJiraSettingsFlag){isJiraConfiguredTimeTrackingEnabled,workingHoursPerDay,workingDaysPerWeek,defaultFormat,defaultUnit}\n    issueLinkTypes(cloudId:$cloudId)@include(if:$giraAGGMigrationJiraSettingsFlag)@optIn(to:\"JiraIssueLinkTypes\"){edges{node{linkTypeId,inwards,outwards,id}}},issueByKey(key:$issueKey,cloudId:$cloudId){screenId,errorRetrievingData,issueId,key,isArchived@include(if:$giraAGGMigrationIsArchivedFlag),hierarchyLevelAbove{level,name},hierarchyLevelBelow{level,name},issueTypesForHierarchyAbove{edges{node{id,issueTypeId,name,description,avatar{xsmall,small,medium,large},hierarchy{level,name}}}},issueTypesForHierarchyBelow{edges{node{id,issueTypeId,name,description,avatar{xsmall,small,medium,large},hierarchy{level,name}}}},issueTypesForHierarchySame{edges{node{id,issueTypeId,name,description,avatar{xsmall,small,medium,large},hierarchy{level,name}}}},...prefetchFieldSuggestions_issueView@include(if:$issueViewRelayLeaderFlag),...src_issueViewLayoutTemplatesItemList_ItemList_4GhV3f,fields{edges{node{__typename,...on JiraAffectedServicesField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedAffectedServicesConnection{edges{node{serviceId,name}}},searchUrl},...on JiraAssetField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedAssetsConnection{edges{node{appKey,originId,serializedOrigin,value}}},searchUrl},...on JiraBooleanField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,value},...on JiraCMDBField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,searchUrl,isInsightAvailable,wasInsightRequestSuccessful,cmdbFieldConfig{attributesDisplayedOnIssue{edges{node}},attributesIncludedInAutoCompleteSearch{edges{node}},objectSchemaId,multiple},selectedCmdbObjectsConnection{edges{node{label,objectGlobalId,objectId,objectKey,workspaceId,webUrl,avatar{mediaClientConfig{clientId,fileId,issuer,mediaBaseUrl,mediaJwtToken},url48},objectType{objectTypeId,name,objectSchemaId,icon{id,name,url16,url48}},attributes{edges{node{attributeId,objectTypeAttributeId,objectTypeAttribute{label,name,defaultType{id,name},type},objectAttributeValues{edges{node{displayValue,searchValue,value,user{__typename,accountId,name,picture,id},status{category,id,name},group{name,id},project{id,name,key,avatar{large}},bitbucketRepo{uuid,name,url,avatarUrl},opsgenieTeam{name,id,url},referencedObject{objectGlobalId,workspaceId,objectKey,label,avatar{mediaClientConfig{clientId,fileId,issuer,mediaBaseUrl,mediaJwtToken},url48},objectId}}}}}}}}}}},...on JiraCascadingSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,cascadingOption{parentOptionValue{id,optionId,value,isDisabled},childOptionValue{id,optionId,value,isDisabled}},cascadingOptions{edges{node{parentOptionValue{id,optionId,value,isDisabled},childOptionValues{id,optionId,value,isDisabled}}}}},...on JiraCheckboxesField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedOptions{edges{node{id,optionId,value,isDisabled}}},fieldOptions{edges{node{id,optionId,value,isDisabled}}}},...on JiraColorField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,color{id,colorKey}},...on JiraComponentsField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedComponentsConnection{edges{node{id,componentId,name,description}}},components{edges{node{id,componentId,name,description}}}},...on JiraConnectDateTimeField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,dateTime},...on JiraConnectMultipleSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedOptions{edges{node{id,optionId,value,isDisabled}}},fieldOptions{edges{node{id,optionId,value,isDisabled}}}},...on JiraConnectNumberField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,number},...on JiraConnectRichTextField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,richText{plainText,adfValue{json}},renderer},...on JiraConnectSingleSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,fieldOption{id,optionId,value,isDisabled},fieldOptions{edges{node{id,optionId,value,isDisabled}}},searchUrl},...on JiraConnectTextField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,text},...on JiraDatePickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,date},...on JiraDateTimePickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,dateTime},...on JiraEpicLinkField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,epic{id,issueId,name,key,summary,color,done},searchUrl},...on JiraFlagField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,flag{isFlagged}},...on JiraForgeGroupField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedGroup{id,groupId,name},renderer},...on JiraForgeGroupsField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedGroupsConnection{edges{node{id,groupId,name}}},renderer},...on JiraForgeNumberField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,number,renderer},...on JiraForgeObjectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,object,renderer},...on JiraForgeStringField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,text,renderer},...on JiraForgeStringsField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedLabelsConnection{edges{node{labelId,name}}},renderer,searchUrl},...on JiraForgeUserField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,user{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{email,zoneinfo,locale},...on CustomerUser{email,zoneinfo,locale},id},renderer,searchUrl},...on JiraForgeUsersField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedUsersConnection{edges{node{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{email,zoneinfo,locale},...on CustomerUser{email,zoneinfo,locale},id}}},renderer,searchUrl},...on JiraIssueRestrictionField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedRolesConnection{edges{node{id,roleId}}},searchUrl},...on JiraIssueTypeField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,issueType{id,issueTypeId,name,description,avatar{xsmall,small,medium,large},hierarchy{level,name}}},...on JiraLabelsField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedLabelsConnection{edges{node{labelId,name}}},searchUrl},...on JiraMultipleGroupPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedGroupsConnection{edges{node{groupId,name,id}}},searchUrl},...on JiraMultipleSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedOptions{edges{node{id,optionId,value,isDisabled}}},fieldOptions{edges{node{id,optionId,value,isDisabled}}},searchUrl},...on JiraMultipleSelectUserPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedUsersConnection{edges{node{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{email,zoneinfo,locale},...on CustomerUser{email,zoneinfo,locale},id}}},searchUrl},...on JiraMultipleVersionPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedVersionsConnection{edges{node{id,versionId,name,status,description}}},versions{edges{node{id,versionId,name,status,description}}}},...on JiraNumberField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,number,isStoryPointField},...on JiraOriginalTimeEstimateField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,originalEstimate{timeInSeconds}},...on JiraParentIssueField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,parentVisibility{hasEpicLinkFieldDependency,canUseParentLinkField},parentIssue{id,issueId,key,webUrl,fieldsById(ids:[\"issuetype\",\"project\",\"summary\"]){edges{node{__typename,...on JiraIssueTypeField{fieldId,name,type,description,issueType{id,issueTypeId,name,description,avatar{xsmall,small,medium,large},hierarchy{level,name}}},...on JiraProjectField{fieldId,name,type,description,searchUrl,project{key,name,projectId,canSetIssueRestriction,projectStyle,projectType,status,avatar{medium,large},similarIssues{featureEnabled},id}},...on JiraSingleLineTextField{fieldId,name,type,description,text},id}}}}},...on JiraPeopleField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedUsersConnection{edges{node{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{email,zoneinfo,locale},...on CustomerUser{email,zoneinfo,locale},id}}},isMulti,searchUrl},...on JiraPriorityField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,priority{priorityId,name,iconUrl,color,id},priorities{edges{node{priorityId,name,iconUrl,color,id}}}},...on JiraProformaFormsField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,proformaForms{hasProjectForms,hasIssueForms,isHarmonisationEnabled}},...on JiraProjectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,searchUrl,project{key,name,projectId,canSetIssueRestriction,projectStyle,projectType,status,avatar{medium,large},similarIssues{featureEnabled},id}},...on JiraRadioSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedOption{id,optionId,value,isDisabled},fieldOptions{edges{node{id,optionId,value,isDisabled}}}},...on JiraResolutionField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,resolution{id,resolutionId,name},resolutions{edges{node{id,resolutionId,name}}}},...on JiraRichTextField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,richText{plainText,adfValue{json}},renderer},...on JiraSecurityLevelField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,securityLevel{name,securityId,description,id},securityLevels{edges{node{name,securityId,description,id}}}},...on JiraServiceManagementApprovalField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,activeApproval{id,approvalState,approverPrincipals{edges{node{__typename,...on JiraServiceManagementUserApproverPrincipal{user{__typename,accountId,name,id},jiraRest},...on JiraServiceManagementGroupApproverPrincipal{groupId,name,memberCount,approvedCount}}}},approvers{edges{node{approver{__typename,accountId,accountStatus,name,...on AtlassianAccountUser{email,zoneinfo},...on CustomerUser{email,zoneinfo},id},approverDecision}}},canAnswerApproval,configurations{approversConfigurations{type,fieldName,fieldId},condition{type,value}},createdDate,decisions{edges{node{approver{__typename,accountId,accountStatus,name,...on AtlassianAccountUser{email,zoneinfo},...on CustomerUser{email,zoneinfo},id},approverDecision}}},excludedApprovers{edges{node{__typename,accountId,accountStatus,name,...on AtlassianAccountUser{email,zoneinfo},...on CustomerUser{email,zoneinfo},id}}},finalDecision,name,pendingApprovalCount,status{id,name,categoryId}},completedApprovalsConnection{edges{node{id,name,finalDecision,approvers{edges{node{approver{__typename,accountId,accountStatus,name,...on AtlassianAccountUser{email,zoneinfo},...on CustomerUser{email,zoneinfo},id},approverDecision}}},createdDate,completedDate,status{name,categoryId}}}}},...on JiraServiceManagementMajorIncidentField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,majorIncident},...on JiraServiceManagementOrganizationField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedOrganizationsConnection{edges{node{organizationId,organizationName,domain}}},searchUrl},...on JiraServiceManagementPeopleField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedUsersConnection{edges{node{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{email,zoneinfo,locale},...on CustomerUser{email,zoneinfo,locale},id}}},searchUrl},...on JiraServiceManagementRequestFeedbackField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,feedback{rating}},...on JiraServiceManagementRequestLanguageField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,language{languageCode,displayName},languages{languageCode,displayName}},...on JiraServiceManagementRequestTypeField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,requestType{id,requestTypeId,name,issueType{id,issueTypeId},avatar{xsmall,small,medium,large}}},...on JiraServiceManagementRespondersField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,searchUrl,respondersConnection{edges{node{__typename,...on JiraServiceManagementUserResponder{user{__typename,canonicalAccountId,picture,name,id}},...on JiraServiceManagementTeamResponder{teamId,teamName}}}}},...on JiraSingleGroupPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedGroup{id,groupId,name},searchUrl},...on JiraSingleLineTextField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,text},...on JiraSingleSelectField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,fieldOption{id,optionId,value,isDisabled},fieldOptions{edges{node{id,optionId,value,isDisabled}}},searchUrl},...on JiraSingleSelectUserPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,user{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{zoneinfo,locale},...on CustomerUser{zoneinfo,locale},id},searchUrl},...on JiraSingleVersionPickerField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,version{id,versionId,name,iconUrl,status,description,startDate,releaseDate},versions{edges{node{id,versionId,name,iconUrl,status,description,startDate,releaseDate}}}},...on JiraSprintField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedSprintsConnection{edges{node{id,sprintId,name,state,boardName,startDate,endDate,completionDate,goal}}},searchUrl},...on JiraStatusField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,status{id,name,description,statusId,statusCategory{id,statusCategoryId,key,name,colorName}}},...on JiraTeamField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedTeam{teamId,name,isShared,avatar{medium},id}},...on JiraTeamViewField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,selectedTeam{jiraSuppliedVisibility,jiraSuppliedName,jiraSuppliedId,jiraSuppliedTeamId,jiraSuppliedAvatar{xsmall,small,medium,large}},searchUrl},...on JiraTimeTrackingField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,originalEstimate{timeInSeconds},remainingEstimate{timeInSeconds},timeSpent{timeInSeconds}},...on JiraUrlField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,uri},...on JiraVotesField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,vote{hasVoted,count}},...on JiraWatchesField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,watch{isWatching,count}},...on JiraForgeDatetimeField{...on JiraIssueField{__isJiraIssueField:__typename,...on JiraIssueFieldConfiguration{__isJiraIssueFieldConfiguration:__typename,fieldConfig{isRequired,isEditable}}},fieldId,name,type,description,dateTime,renderer},id}}},childIssues{__typename,...on JiraChildIssuesWithinLimit{issues(activeProjectsOnly:true){edges{node{key,issueId,webUrl,storyPointEstimateField{__typename,fieldId,type,number,isStoryPointField,id},storyPointsField{__typename,fieldId,type,number,isStoryPointField,id},fieldsById(ids:[\"assignee\",\"created\",\"issuetype\",\"priority\",\"status\",\"summary\",\"timetracking\",\"updated\"]){edges{node{__typename,...on JiraIssueTypeField{fieldId,type,issueType{issueTypeId,name,avatar{medium},hierarchy{level},id}},...on JiraSingleLineTextField{fieldId,type,text},...on JiraPriorityField{fieldId,type,priority{priorityId,name,iconUrl,id}},...on JiraStatusField{fieldId,type,status{name,statusId,statusCategory{statusCategoryId,id},id}},...on JiraSingleSelectUserPickerField{fieldId,type,user{__typename,accountId,accountStatus,name,picture,id}},...on JiraTimeTrackingField{fieldId,type,originalEstimate{timeInSeconds},remainingEstimate{timeInSeconds},timeSpent{timeInSeconds}},...on JiraDateTimePickerField{fieldId,type,dateTime},id}}},id}}}},...on JiraChildIssuesExceedingLimit{search}},issueLinks(first:1000){edges{node{id,issueLinkId,direction,relationName,type{linkTypeId,id},issue{id,issueId,key,webUrl,fieldsById(ids:[\"assignee\",\"issuetype\",\"priority\",\"status\",\"summary\"]){edges{node{__typename,...on JiraStatusField{fieldId,name,type,status{id,name,statusId,statusCategory{id,statusCategoryId,key,name,colorName}}},...on JiraPriorityField{fieldId,name,type,priority{priorityId,name,iconUrl,color,id}},...on JiraIssueTypeField{fieldId,name,type,issueType{id,issueTypeId,name,avatar{xsmall,small,medium,large},hierarchy{level,name}}},...on JiraSingleLineTextField{fieldId,name,type,text},...on JiraSingleSelectUserPickerField{fieldId,name,type,user{__typename,accountId,accountStatus,name,picture,...on AtlassianAccountUser{zoneinfo,locale},...on CustomerUser{zoneinfo,locale},id}},id}}}}}}},id}}}fragment issueViewLabelsSystemField_issueViewLayoutTemplatesLayoutItem on JiraLabelsField{...labels_issueFieldLabelsInlineEditViewFull_LabelsInlineEditView,name,description,fieldId,selectedLabelsConnection(first:100){totalCount},fieldConfig{isEditable}}fragment labels_issueFieldLabelsEditviewFull_LabelsEditViewWithSuggestionsFragment_nodeIdFragment on JiraLabelsField{id,fieldId}fragment labels_issueFieldLabelsEditviewFull_LabelsEditView_nodeIdFragment on JiraLabelsField{...labels_issueFieldLabelsEditviewFull_LabelsEditViewWithSuggestionsFragment_nodeIdFragment,id}fragment labels_issueFieldLabelsInlineEditViewFull_LabelsInlineEditView on JiraLabelsField{...labels_issueFieldLabelsReadviewFull_LabelsReadView,...labels_issueFieldLabelsEditviewFull_LabelsEditView_nodeIdFragment,id,name,fieldId,type,selectedLabelsConnection(first:100){totalCount,edges{node{name}}},fieldConfig{isEditable}}fragment labels_issueFieldLabelsReadviewFull_LabelsReadView on JiraLabelsField{fieldId,selectedLabelsConnection(first:100){totalCount,edges{node{name}}}}fragment prefetchFieldSuggestions_issueView on JiraIssue{prefetchFieldIds:fieldsById(ids:[\"labels\"]){edges{node{__typename,id,fieldId}}}}fragment relayLayoutItems_issueViewLayoutTemplatesLayoutItem_RelayLayoutItemFragmentContainer_3ciNzv on Node{__isNode:__typename,...issueViewLabelsSystemField_issueViewLayoutTemplatesLayoutItem@include(if:$issueViewRelayLabelsFlag)}fragment src_issueViewLayoutTemplatesItemList_ItemList_4GhV3f on JiraIssue{fieldFragments:fields@include(if:$issueViewRelayLeaderFlag){edges{node{__typename,__isJiraIssueField:__typename,fieldId,...src_issueViewLayoutTemplatesLayoutItem_LayoutItemInternal_3ciNzv,id}}}}fragment src_issueViewLayoutTemplatesLayoutItem_LayoutItemInternal_3ciNzv on Node{__isNode:__typename,__typename,...on JiraIssueField{__isJiraIssueField:__typename,type},...relayLayoutItems_issueViewLayoutTemplatesLayoutItem_RelayLayoutItemFragmentContainer_3ciNzv}
  """
variables = {
  'cloudId':"ec6b7b36-879d-44f1-a768-6adad0c38ed5",
  'giraAGGMigrationIsArchivedFlag': False,
  'giraAGGMigrationJiraSettingsFlag': False,
  'issueKey':"MWS-136550",
  'issueViewRelayLabelsFlag': False,
  'issueViewRelayLeaderFlag': False
}

#for writing into excel file
wb = openpyxl.Workbook()

ws = wb.active

ws.title = "Updated log"

excel_col = 1
ws.cell(row = 1, column = 1, value = "issue_key")
ws.cell(row = 1, column = 2, value = "Content")
ws.cell(row = 1, column = 3, value = "issue_type")

#for ChatGPT
API_KEY = "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"
API_ENDPOINT = "https://api.openai.com/v1/chat/completions"

def generate_chat_completion(messages, model="gpt-4", temperature=1, max_tokens=None):
    gpt_headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}",
    }

    gpt_data = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
    }

    if max_tokens is not None:
        gpt_data["max_tokens"] = max_tokens

    response2 = requests.post(API_ENDPOINT, headers=gpt_headers, data=json.dumps(gpt_data))

    if response2.status_code == 200:
        return response2.json()["choices"][0]["message"]["content"]
    else:
        raise Exception(f"Error {response2.status_code}: {response2.text}")

#main process
# issue_keys = ["MWS-136939"]
for ind in issue_keys:
  variables['issueKey'] = ind

  if ind == "MWS-136240" or ind == "MWS-136274" or ind == "MWS-131970" or ind == "MWS-131977" or ind == "MWS-130888" or ind == "MWS-121491":
      continue

  print(ind)
  response = requests.request(
    "POST",
    url,
    headers=headers,
    auth=auth,
    json={'query': query, 'variables': variables}
  )

  # with open("output.txt", "w") as file:
  #     file.write(response.text)

  # print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))

  data = json.loads(response.text)
  messageContent = """
    These are cagtegories.
    1. Task - All MWS tickets are automatically assigned to this category by default. The issue should only be sorted into this category if it does not fit into any other existing categories.
    
    2. Account Edit/Creation - Whenever a request is submitted for the creation of an MPN or Customer Portal account, or when a change needs to be made to an existing account.
    
    3. Account Login - A customer with an ‘MPN' or 'Customer Portal’ account is unable to log in. Although an account exists, it cannot be accessed.
    
    4. Analyst Meeting Needed - In order to better understand the jurisdiction's overall functionality needs, it is necessary to meet with Tier II or senior analysts. This meeting will allow us to provide recommendations regarding configuration and other aspects of the system.
    
    5. API - Changes or requests related to APIs.
    
    6. Application Creation/Edits - The creation or editing of applications. "New Implementation/Major Configuration" should be used for modifications that do not make sense or are likely to have a significant impact. "Analysis Meeting Needed" should be used if additional information or a more detailed discussion is required.
    
    7. Comment Letter Comments - Edit/create 'Comments' or 'Comment Categories'. "Comment Letter Comments" are also referred to as "Review Comments".
    
    8. Contractor Licensing - Anything related to contractor licensing and renewals.
    
    9. Customer Portal (Mobile) - All matters relating to the customer portal mobile apps (MGO Connect for iOS or Google Play).
    
    10. Customer Portal (Web) - Any questions regarding the web versions of the customer portal (MGO CP  or MGO Connect). This includes general customer questions.
    
    11. Data Import/Export - Data import or export requests that are not handled via an API or report.
    
    12. DB Config Needed - We use this category when an item cannot be configured via the UI interface 'MPN' and must be managed through the database. Note: this is not a development item, but rather something we already know how to accomplish through database configuration.
    
    13. Fees - Whenever fees need to be changed or added. This category should be the default when sorting fee-related tickets so that it can be reviewed further by someone with expertise in the area.
    
    14. Fees: Database - The configuration of fees that may only be accomplished via the database. Only the Fees team should utilize this category.
    
    15. Fees: GL Codes - Adding or updating general ledger codes. Only the Fees team should utilize this category.
    
    16. GIS Integration - GIS issues, such as changes to GIS integration, GIS errors, etc. GIS can also be referred to as 'Verify Address'.
    
    17. GPS - Anything relating to GPS functionality. GPS can also be referred to as 'Asset Tracking'. GPS should not be confused with GIS.
    
    18. MGO Connect Live - Anything pertaining to MGO Connect Live Conferences, including attendance and registration.
    
    19. MGO Tasks General - Any issue related to the MGO Tasks app that is not a work order. Use "WO (IA, Tasks, MGOC, Board)" if there is a problem with the work order/inspection functionality.
    
    20. New Feature - Essentially, this is a request for a feature that cannot be accomplished with the current capabilities of the software. For example, "we would like the software to prepare coffee for us. It does not currently do so, but we would really like it if it did."
    
    21. New Implementation/Major Config -The implementation of a new project type after it has been sold to a jurisdiction, or when the jurisdiction wishes to make significant changes to its existing configuration.
    
    22. Payment Integration - This category should only be used if an integration issue is present that affects the customer portal or in-person credit cards. If it is simply a missing payment that needs to be resolved, it should be sorted under "Fees".
    
    23. Project Custom Fields - When a 'Project Custom Field' needs to be created or updated. It is imperative to note that "Project Custom Fields" is also referred to as "Project Checklists".
    
    24. Project Status Update - Add a 'Status' to the drop-down menu on the project page or change the title of a Status. This may also apply to what actions are available in the customer portal during a Status, such as uploading files, searching for projects, etc.
    
    25. Reports/Letters - The creation and editing of Reports, Form Letters, or other SSRS documents.
    
    26. Requirement & Inspection Creation/Edits - The creation or editing of 'Requirement Sets', 'Requirements', and 'Inspections' that are not complex and do not require extensive business analysis. Additionally, the jurisdiction should justify the change and understand its impact. "Analysis Meeting Needed" should be used when additional information is required, a more in-depth discussion is required, the request is not reasonable, or the request is likely to significantly impact the configuration of the jurisdiction. 
    
    27. Sales - Anything related to sales and public relations. For instance, we are interested in permit software. Can your CTO contact me? Does your software accomplish this? What is the SCPDC?
    
    28. Sub-task - The 'Create Subtask' button near the top of a ticket triggers this category. This feature should only be used when an MWS ticket contains multiple issues that require separate attention.
    
    29. Tier II - This category is no longer in use.
    
    30. Utility Notification - Anything related to 'Utility Notifications'.
    
    31. Virtual Inspections - Topics related to virtual inspections in MGO Connect.
    
    32. WO (IA, Tasks, MGOC, Board) - Any issue related to Work Order or Inspection functionality that appears in 'Inspection Anywhere', 'MGO Tasks App' (WO portion only), 'MGO Connect Web Version' (WO functionality only), or the 'Work Order Management Board'."

    This is the text.
  """
  array1 = data['data']['jira']['issueByKey']['fields']['edges'][7]['node']['richText']['adfValue']['json']['content']

  mainMessage = "Description:"
  for i in array1:
      if i['type'] == "paragraph":
        for j in i['content']:
            if j['type'] == "text":
                mainMessage += j['text']
                mainMessage += "\n"
  
  messageContent += mainMessage
  messageContent += "\nTell me the most fit category of above text in a word. If it is difficult to answer in a word, just tell me any category randomly."
  messages = [
    # {"role": "system", "content": "You are a helpful assistant."},
    {"role": "user", "content": messageContent}
  ]
  # print(messageContent)

  response_text = generate_chat_completion(messages)
  print(ind)
  print("-MainMessage : \n", mainMessage)
  print("*** According to the ChatGPT answer ***")
  print(response_text)
  print()
  print()
  
  excel_col += 1
  cell = ws.cell(row=excel_col, column=1)
  cell.value = ind
  ws.cell(row = excel_col, column = 2, value = mainMessage)
  ws.cell(row = excel_col, column = 3, value = response_text)

  wb.save("test.xlsx")